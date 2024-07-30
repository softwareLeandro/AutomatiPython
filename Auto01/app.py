# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from time import sleep
# import openpyxl

# # Caminho para o seu perfil específico do Chrome
# chrome_profile_path = "/Users/leandroalexandre/Library/Application Support/Google/Chrome/Default"

# # Configurar opções do Chrome para usar o perfil específico
# options = webdriver.ChromeOptions()
# options.add_argument(f"user-data-dir={chrome_profile_path}")

# # Inicializar o WebDriver com as opções configuradas
# driver = webdriver.Chrome(options=options)

# # Navegar para a URL
# driver.get('https://contabilidade-devaprender.netlify.app/')
# sleep(3)

# # Localizar e preencher o campo de email
# email = driver.find_element(By.XPATH, "//input[@id='email']")
# sleep(2)
# email.send_keys('alexandreleandrodev@gmail.com')

# # Localizar e preencher o campo de senha
# senha = driver.find_element(By.XPATH, "//input[@id='senha']")
# sleep(2)
# senha.send_keys('Chica12345')

# # Localizar e clicar no botão entrar
# botao_entrar = driver.find_element(By.XPATH, "//button[@id='Entrar']")
# sleep(2)
# botao_entrar.click()

# empresas = openpyxl.load_workbook('./empresas.xlsx')
# paginas_empresas = empresas['dados empresas']

# for linha in paginas_empresas.iter_rows(min_row=2, values_only=True):
#     nome_empresa, email, telefone, endereco, cnpj, area_atuacao, quantidade_de_funcionarios, data_fundacao = linha

# driver.find_element(By.ID, 'nomeEmpresa').send_keys(nome_empresa)
# sleep(1)
# driver.find_element(By.ID, 'emailEmpresa').send_keys(email)
# sleep(1)
# driver.find_element(By.ID, 'telefoneEmpresa').send_keys(telefone)
# sleep(1)
# driver.find_element(By.ID, 'enderecoEmpresa').send_keys(endereco)
# sleep(1)
# driver.find_element(By.ID, 'cnpj').send_keys(cnpj)
# sleep(1)
# driver.find_element(By.ID, 'areaAtuacao').send_keys(area_atuacao)
# sleep(1)
# driver.find_element(By.ID, 'numeroFuncionarios').send_keys(
#     quantidade_de_funcionarios)
# sleep(1)
# driver.find_element(By.ID, 'dataFundacao').send_keys(data_fundacao)
# sleep(1)

# driver.find_element(By.ID, 'Cadastrar').click()
# sleep(3)


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
import openpyxl

# Caminho para o seu perfil específico do Chrome
chrome_profile_path = "/Users/leandroalexandre/Library/Application Support/Google/Chrome/Default"

# Configurar opções do Chrome para usar o perfil específico
options = webdriver.ChromeOptions()
options.add_argument(f"user-data-dir={chrome_profile_path}")

# Inicializar o WebDriver com as opções configuradas
driver = webdriver.Chrome(options=options)

# Navegar para a URL
driver.get('https://contabilidade-devaprender.netlify.app/')

# Esperar que o campo de email esteja presente
wait = WebDriverWait(driver, 10)
email = wait.until(EC.presence_of_element_located(
    (By.XPATH, "//input[@id='email']")))
email.send_keys('alexandreleandrodev@gmail.com')

# Localizar e preencher o campo de senha
senha = driver.find_element(By.XPATH, "//input[@id='senha']")
senha.send_keys('Chica12345')

# Localizar e clicar no botão entrar
botao_entrar = driver.find_element(By.XPATH, "//button[@id='Entrar']")
botao_entrar.click()

# Carregar a planilha de Excel
empresas = openpyxl.load_workbook('./empresas.xlsx')
paginas_empresas = empresas['dados empresas']

# Função para preencher um campo com atraso


def preencher_campo(elemento, valor):
    for caractere in valor:
        elemento.send_keys(caractere)
        sleep(0.1)  # Ajuste o tempo de atraso conforme necessário


# Iterar pelas linhas da planilha
for linha in paginas_empresas.iter_rows(min_row=2, values_only=True):
    nome_empresa, email, telefone, endereco, cnpj, area_atuacao, quantidade_de_funcionarios, data_fundacao = linha

    # Preencher os campos no formulário
    nome_empresa_elemento = wait.until(
        EC.presence_of_element_located((By.ID, 'nomeEmpresa')))
    preencher_campo(nome_empresa_elemento, nome_empresa)

    email_empresa_elemento = driver.find_element(By.ID, 'emailEmpresa')
    preencher_campo(email_empresa_elemento, email)

    telefone_empresa_elemento = driver.find_element(By.ID, 'telefoneEmpresa')
    preencher_campo(telefone_empresa_elemento, telefone)

    endereco_empresa_elemento = driver.find_element(By.ID, 'enderecoEmpresa')
    preencher_campo(endereco_empresa_elemento, endereco)

    cnpj_elemento = driver.find_element(By.ID, 'cnpj')
    preencher_campo(cnpj_elemento, cnpj)

    area_atuacao_elemento = driver.find_element(By.ID, 'areaAtuacao')
    preencher_campo(area_atuacao_elemento, area_atuacao)

    numero_funcionarios_elemento = driver.find_element(
        By.ID, 'numeroFuncionarios')
    preencher_campo(numero_funcionarios_elemento,
                    str(quantidade_de_funcionarios))

    data_fundacao_elemento = driver.find_element(By.ID, 'dataFundacao')
    preencher_campo(data_fundacao_elemento, data_fundacao)

    # Clicar no botão cadastrar
    driver.find_element(By.ID, 'Cadastrar').click()
    sleep(3)  # Ajuste o tempo de espera conforme necessário

# Fechar o navegador
driver.quit()
